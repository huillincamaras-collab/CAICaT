# gui_auditor.py
"""
Módulo de auditoría mejorado con soporte para corrección de etiquetados.
"""

import tkinter as tk
from tkinter import messagebox, simpledialog, Text, Scrollbar
from PIL import Image, ImageTk
import cv2
import os
import json
import threading
import subprocess
from datetime import datetime
from config_utils import metadata_lock, load_config
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


def open_file_default(path):
    """Abre un archivo con la aplicación predeterminada del sistema."""
    if os.name == "nt":
        os.startfile(path)
    elif os.uname().sysname == "Darwin":
        subprocess.Popen(["open", path])
    else:
        subprocess.Popen(["xdg-open", path])


class AuditorGUI(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("🔍 Auditoría - CAICAT v2.2")
        self.geometry("1000x750")
        self.configure(bg="#f5f5f5")
        
        self.config_data = load_config()
        self.output_folder = self.config_data.get("General", {}).get("output_folder", "output")
        
        # ✨ SOLICITAR NOMBRE DEL AUDITOR
        self.auditor_name = simpledialog.askstring(
            "Auditoría",
            "Ingrese su nombre (auditor):",
            parent=self
        )
        
        if not self.auditor_name:
            messagebox.showwarning("Sin auditor", "No se ingresó nombre de auditor. Usando 'Anónimo'.")
            self.auditor_name = "Anónimo"
        
        # Detectar última sesión
        self.metadata_path = self._auto_detect_last_session()
        if not self.metadata_path:
            messagebox.showwarning("Auditoría", "No se encontraron sesiones procesadas.")
            self.destroy()
            return
        
        self.video_dirs = []
        self.current_idx = 0
        self.tk_img = None
        
        self.load_metadata()
        self.build_layout()
        
        # Atajos de teclado
        self.bind("<Left>", lambda e: self.prev_video())
        self.bind("<Right>", lambda e: self.next_video())
        self.bind("<Control-s>", lambda e: self.export_audit_report())
        
        self.after(100, self.show_frame)

    def _auto_detect_last_session(self):
        """Detecta la última sesión procesada."""
        sessions_dir = os.path.join(self.output_folder, "sessions")
        if not os.path.exists(sessions_dir):
            return None
        
        latest = None
        max_mtime = 0
        
        for item in os.listdir(sessions_dir):
            metadata_path = os.path.join(sessions_dir, item, "metadata.json")
            if os.path.exists(metadata_path):
                mtime = os.path.getmtime(metadata_path)
                if mtime > max_mtime:
                    max_mtime = mtime
                    latest = metadata_path
        
        return latest

    def load_metadata(self):
        """Carga metadata e inicializa campos de auditoría."""
        with open(self.metadata_path, "r", encoding="utf-8") as f:
            self.video_dirs = json.load(f)
        
        # Inicializar campos de auditoría si no existen
        for v in self.video_dirs:
            if "audit" not in v:
                v["audit"] = {
                    "status": "pending",
                    "auditor_name": "",
                    "audit_date": "",
                    "correction_notes": "",
                    "requires_retagging": False,
                    "retagged": False
                }

    def build_layout(self):
        """Construye la interfaz de auditoría."""
        # Header
        header = tk.Frame(self, bg="#2196f3", height=50)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        tk.Label(
            header,
            text=f"🔍 Auditoría - {self.auditor_name}",
            bg="#2196f3",
            fg="white",
            font=("Arial", 14, "bold")
        ).pack(pady=12)
        
        # Navegación
        nav_frame = tk.Frame(self, bg="#f5f5f5")
        nav_frame.pack(fill="x", pady=5)
        
        tk.Button(
            nav_frame,
            text="◀ Anterior",
            command=self.prev_video,
            width=10
        ).pack(side="left", padx=5)
        
        self.video_label = tk.Label(
            nav_frame,
            text="",
            font=("Arial", 11),
            bg="#f5f5f5"
        )
        self.video_label.pack(side="left", fill="x", expand=True, padx=10)
        
        self.status_label = tk.Label(
            nav_frame,
            text="",
            font=("Arial", 10, "bold"),
            fg="gray"
        )
        self.status_label.pack(side="right", padx=10)
        
        tk.Button(
            nav_frame,
            text="Siguiente ▶",
            command=self.next_video,
            width=10
        ).pack(side="right", padx=5)
        
        # Canvas para imagen
        self.canvas = tk.Canvas(self, width=912, height=513, bg="black", bd=2, relief="sunken")
        self.canvas.pack(pady=10)
        
        # Frame de información del video
        info_frame = tk.LabelFrame(self, text="Información del Video", bg="#f5f5f5", font=("Arial", 10, "bold"))
        info_frame.pack(fill="x", padx=10, pady=5)
        
        self.info_text = tk.Label(
            info_frame,
            text="",
            bg="#f5f5f5",
            font=("Arial", 9),
            justify="left",
            anchor="w"
        )
        self.info_text.pack(fill="x", padx=10, pady=5)
        
        # ✨ TextBox para notas de corrección
        notes_frame = tk.LabelFrame(self, text="Notas de Corrección", bg="#f5f5f5", font=("Arial", 10, "bold"))
        notes_frame.pack(fill="x", padx=10, pady=5)
        
        text_container = tk.Frame(notes_frame, bg="#f5f5f5")
        text_container.pack(fill="x", padx=5, pady=5)
        
        self.notes_text = Text(text_container, height=3, width=80, font=("Arial", 9))
        self.notes_text.pack(side="left", fill="x", expand=True)
        
        scrollbar = Scrollbar(text_container, command=self.notes_text.yview)
        scrollbar.pack(side="right", fill="y")
        self.notes_text.config(yscrollcommand=scrollbar.set)
        
        # ✨ Checkbox "Requiere re-etiquetado"
        self.requires_retagging_var = tk.BooleanVar(value=False)
        tk.Checkbutton(
            notes_frame,
            text="⚠️ Requiere re-etiquetado",
            variable=self.requires_retagging_var,
            bg="#f5f5f5",
            font=("Arial", 10, "bold"),
            fg="#FF5722"
        ).pack(anchor="w", padx=10, pady=5)
        
        # Botones de acción
        btn_frame = tk.Frame(self, bg="#f5f5f5")
        btn_frame.pack(pady=10)
        
        tk.Button(
            btn_frame,
            text="✅ Aprobar",
            command=self.approve,
            bg="#4CAF50",
            fg="white",
            width=12,
            height=2,
            font=("Arial", 10)
        ).pack(side="left", padx=10)
        
        tk.Button(
            btn_frame,
            text="⚠️ Marcar Revisar",
            command=self.flag_review,
            bg="#FF9800",
            fg="white",
            width=14,
            height=2,
            font=("Arial", 10)
        ).pack(side="left", padx=10)
        
        tk.Button(
            btn_frame,
            text="🎬 Abrir Video",
            command=self.open_original,
            bg="#607D8B",
            fg="white",
            width=12,
            height=2,
            font=("Arial", 10)
        ).pack(side="left", padx=10)
        
        tk.Button(
            btn_frame,
            text="📊 Exportar Reporte",
            command=self.export_audit_report,
            bg="#9C27B0",
            fg="white",
            width=14,
            height=2,
            font=("Arial", 10)
        ).pack(side="left", padx=10)
        
        # Footer
        footer = tk.Frame(self, bg="#f0f0f0", height=30)
        footer.pack(fill="x", side="bottom")
        footer.pack_propagate(False)
        
        tk.Label(
            footer,
            text="Atajos: ← → Navegar | Ctrl+S Exportar reporte",
            bg="#f0f0f0",
            font=("Arial", 8),
            fg="gray"
        ).pack(pady=5)

    def show_frame(self):
        """Muestra el frame del video actual."""
        if not self.video_dirs:
            self.canvas.create_text(450, 250, text="Sin datos", fill="white")
            return

        v = self.video_dirs[self.current_idx]

        # Cargar imagen
        frames = []
        for p in v.get("original_photos", []) + v.get("tops", []):
            if p and os.path.exists(p):
                frames.append(p)

        img_path = frames[0] if frames else None

        if img_path:
            img = cv2.imread(img_path)
            if img is not None:
                img = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
                pil = Image.fromarray(img)
                pil.thumbnail((912, 513), Image.Resampling.LANCZOS)
                self.tk_img = ImageTk.PhotoImage(pil)
                self.canvas.delete("all")
                self.canvas.create_image(456, 256, anchor="center", image=self.tk_img)
            else:
                self.canvas.delete("all")
                self.canvas.create_text(450, 250, text="Error al leer frame", fill="red")
        else:
            self.canvas.delete("all")
            self.canvas.create_text(450, 250, text="No hay frames", fill="yellow")

        # ✅ CAMBIO: Fallback file.video_path → video_path
        name = os.path.basename(v.get("file", {}).get("video_path") or v.get("video_path", ""))
        self.video_label.config(text=f"Video {self.current_idx+1}/{len(self.video_dirs)}: {name}")

        # Estado de auditoría
        audit = v.get("audit", {})
        status = audit.get("status", "pending")
        status_map = {
            "pending": "🔹 Pendiente",
            "approved": "✅ Aprobado",
            "needs_correction": "⚠️ Requiere Corrección"
        }
        self.status_label.config(text=status_map.get(status, "🔹 Pendiente"))

        # Información del video
        metadata = v.get("metadata", {})
        classification = v.get("classification", {})
        species = classification.get("species", [])
        counts = classification.get("counts", {})
        behaviors = classification.get("behaviors", [])

        info = f"""
            Site: {metadata.get('site', 'N/A')} | Camera: {metadata.get('camera', 'N/A')} | Operator: {metadata.get('operator', 'N/A')}
            Especies: {', '.join(species) if species else 'Sin etiquetar'}
            Conteos: {', '.join(f'{k}:{v}' for k, v in counts.items()) if counts else 'N/A'}
            Comportamientos: {', '.join(behaviors) if behaviors else 'N/A'}
            """.strip()

        self.info_text.config(text=info)

        # Cargar notas de corrección previas
        self.notes_text.delete("1.0", tk.END)
        correction_notes = audit.get("correction_notes", "")
        if correction_notes:
            self.notes_text.insert("1.0", correction_notes)

        # Cargar checkbox
        self.requires_retagging_var.set(audit.get("requires_retagging", False))

    def _save_audit(self, new_status):
        """Guarda el estado de auditoría del video actual."""
        if not self.video_dirs:
            return
        
        v = self.video_dirs[self.current_idx]
        
        if "audit" not in v:
            v["audit"] = {}
        
        # Actualizar campos de auditoría
        v["audit"]["status"] = new_status
        v["audit"]["auditor_name"] = self.auditor_name
        v["audit"]["audit_date"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        v["audit"]["correction_notes"] = self.notes_text.get("1.0", tk.END).strip()
        v["audit"]["requires_retagging"] = self.requires_retagging_var.get()
        
        # Guardar
        with metadata_lock:
            with open(self.metadata_path, "w", encoding="utf-8") as f:
                json.dump(self.video_dirs, f, indent=4, ensure_ascii=False)
        
        self.show_frame()
        
        # Avanzar al siguiente
        if self.current_idx < len(self.video_dirs) - 1:
            self.next_video()
        else:
            messagebox.showinfo("Auditoría", "🎉 Has revisado el último video de esta sesión.")

    def approve(self):
        """Marca el video como aprobado."""
        if not self.video_dirs:
            return

        v = self.video_dirs[self.current_idx]
        audit = v.get("audit", {})

        # ✅ CAMBIO: Si requería re-etiquetado, marcar como re-etiquetado
        if audit.get("requires_retagging", False):
            v.setdefault("audit", {})["retagged"] = True

        self._save_audit("approved")
        
    def flag_review(self):
        """Marca el video para corrección."""
        self._save_audit("needs_correction")

    def open_original(self):
        """Abre el video original."""
        if self.video_dirs:
            v = self.video_dirs[self.current_idx]
            # ✅ CAMBIO: Fallback file.video_path → video_path
            path = v.get("file", {}).get("video_path") or v.get("video_path", "")
            if path and os.path.exists(path):
                open_file_default(path)
            else:
                messagebox.showwarning("Video no encontrado", "No se pudo encontrar el archivo de video.")

    def next_video(self):
        """Avanza al siguiente video."""
        if self.current_idx < len(self.video_dirs) - 1:
            self.current_idx += 1
            self.show_frame()

    def prev_video(self):
        """Retrocede al video anterior."""
        if self.current_idx > 0:
            self.current_idx -= 1
            self.show_frame()

    def export_audit_report(self):
        """✨ Exporta reporte de auditoría a Excel."""
        try:
            # Crear workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Reporte Auditoría"

            # Estilos
            header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")

            # Headers
            headers = [
                "Video", "Estado", "Auditor", "Fecha Auditoría",
                "Site", "Camera", "Operator",
                "Especies", "Conteos", "Comportamientos",
                "Requiere Re-etiquetado", "Notas de Corrección"
            ]

            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            # Datos
            row = 2
            for v in self.video_dirs:
                audit = v.get("audit", {})
                metadata = v.get("metadata", {})
                classification = v.get("classification", {})

                # ✅ CAMBIO: Fallback file.video_path → video_path
                video_path = v.get("file", {}).get("video_path") or v.get("video_path", "")
                ws.cell(row, 1, os.path.basename(video_path))
                ws.cell(row, 2, audit.get("status", "pending"))
                ws.cell(row, 3, audit.get("auditor_name", ""))
                ws.cell(row, 4, audit.get("audit_date", ""))
                ws.cell(row, 5, metadata.get("site", ""))
                ws.cell(row, 6, metadata.get("camera", ""))
                ws.cell(row, 7, metadata.get("operator", ""))
                ws.cell(row, 8, ", ".join(classification.get("species", [])))
                ws.cell(row, 9, str(classification.get("counts", {})))
                ws.cell(row, 10, ", ".join(classification.get("behaviors", [])))
                ws.cell(row, 11, "Sí" if audit.get("requires_retagging") else "No")
                ws.cell(row, 12, audit.get("correction_notes", ""))
                row += 1

            # Ajustar anchos
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            # Guardar
            output_path = os.path.join(
                os.path.dirname(self.metadata_path),
                f"audit_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            wb.save(output_path)

            messagebox.showinfo(
                "Reporte Exportado",
                f"Reporte guardado en:\n{output_path}"
            )

            # Abrir archivo
            if messagebox.askyesno("Abrir reporte", "¿Desea abrir el reporte ahora?"):
                open_file_default(output_path)

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar el reporte:\n{e}")

if __name__ == "__main__":
    AuditorGUI().mainloop()
